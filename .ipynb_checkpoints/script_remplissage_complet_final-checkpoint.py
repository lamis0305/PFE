
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from difflib import get_close_matches

DOSSIER_SOURCES = "."
FICHIER_TEMPLATE = "centralized_data.xlsx"
FICHIER_LOG = "log_traitement.xlsx"
ANNEES = ["2023"]

def standardiser_nom(nom):
    return str(nom).strip().lower().replace(" ", "").replace("_", "").replace("-", "")

def trouver_compagnie_equivalente(nom_ca, noms_indicateurs):
    match = get_close_matches(standardiser_nom(nom_ca), [standardiser_nom(n) for n in noms_indicateurs], n=1, cutoff=0.8)
    if not match:
        return None
    for original in noms_indicateurs:
        if standardiser_nom(original) == match[0]:
            return original
    return None

def normaliser_nom_branche(nom):
    nom = str(nom).replace("ASS. VIE", "Assurance Vie")
    return "_".join([mot.capitalize() for mot in nom.strip().split()])

def nettoyer_valeur(x):
    if pd.isna(x) or str(x).strip().lower() in ["..", "n.d.", "--", "-", ""]:
        return 0
    try:
        x = str(x).replace(" ", "").replace(" ", "").replace(",", ".")
        return float(x)
    except:
        return 0

def get_id(df, col_nom, valeur, col_id):
    ligne = df[df[col_nom] == valeur]
    return int(ligne[col_id].values[0]) if not ligne.empty else None

if not os.path.exists(FICHIER_TEMPLATE):
    wb = Workbook()
    wb.remove(wb.active)
    for feuille in ["dim_compagnie", "dim_branche", "dim_temps", "Faits_KPI_Assurance"]:
        wb.create_sheet(title=feuille)
    wb.save(FICHIER_TEMPLATE)

if not os.path.exists(FICHIER_LOG):
    pd.DataFrame(columns=["Annee", "NomFichier"]).to_excel(FICHIER_LOG, index=False)

log_df = pd.read_excel(FICHIER_LOG)
wb = load_workbook(FICHIER_TEMPLATE)

for annee in ANNEES:
    chemin = os.path.join(DOSSIER_SOURCES, annee)
    if not os.path.exists(chemin): continue

    fichiers = os.listdir(chemin)
    fichiers_deja_faits = log_df[log_df["Annee"] == int(annee)]["NomFichier"].tolist()

    fichier_CA = [f for f in fichiers if "CHIFFRES" in f.upper()][0]
    fichier_IND = [f for f in fichiers if "PRINCIPAUX" in f.upper()][0]
    fichier_RES = [f for f in fichiers if "RESULTAT TECHNIQUE" in f.upper()][0]
    fichier_SIN = [f for f in fichiers if "SINISTRES REGLES" in f.upper()][0]
    fichier_CPT = [f for f in fichiers if "COMPTE" in f.upper()][0]
    if fichier_IND in fichiers_deja_faits: continue

    df_ca = pd.read_excel(os.path.join(chemin, fichier_CA))
    df_ind = pd.read_excel(os.path.join(chemin, fichier_IND))
    df_res = pd.read_excel(os.path.join(chemin, fichier_RES))
    df_sin = pd.read_excel(os.path.join(chemin, fichier_SIN))
    df_cpt = pd.read_excel(os.path.join(chemin, fichier_CPT))

    def charger_dim(nom):
        colonnes_defaut = {
            "dim_compagnie": ["id_compagnie", "Nom_Compagnie", "Groupe", "Is_Marché"],
            "dim_branche": ["id_branche", "Nom_Branche", "Type_Assurance"],
            "dim_temps": ["id_temps", "Annee"],
            "Faits_KPI_Assurance": ["id_fait", "id_temps", "id_compagnie", "id_branche",
                                    "primes_nettes", "resultat_technique", "sinistres_regles",
                                    "primes_cedees", "provisions_techniques", "resultat_net",
                                    "capitaux_propres", "primes_acquises", "charges_sinistres",
                                    "charges_acquisition_gestion_nettes"]
        }
        data = list(wb[nom].values)
        if not data or all(v is None for v in data[0]):
            return pd.DataFrame(columns=colonnes_defaut[nom])
        return pd.DataFrame(data[1:], columns=data[0])

    dim_compagnie = charger_dim("dim_compagnie")
    dim_branche = charger_dim("dim_branche")
    dim_temps = charger_dim("dim_temps")
    faits = charger_dim("Faits_KPI_Assurance")

    if not (dim_temps["Annee"] == int(annee)).any():
        id_temps = 0 if dim_temps.empty else int(dim_temps["id_temps"].max()) + 1
        dim_temps = pd.concat([dim_temps, pd.DataFrame([{"id_temps": id_temps, "Annee": int(annee)}])], ignore_index=True)
    else:
        id_temps = int(dim_temps[dim_temps["Annee"] == int(annee)]["id_temps"].values[0])

    compagnies = df_ind["COMPAGNIES"].dropna().unique().tolist()
    next_id = 100 if dim_compagnie.empty else int(dim_compagnie["id_compagnie"].max()) + 1
    for nom in compagnies:
        if nom not in dim_compagnie["Nom_Compagnie"].values:
            groupe = nom.replace("VIE", "").strip() if "VIE" in nom else nom
            dim_compagnie = pd.concat([dim_compagnie, pd.DataFrame([{
                "id_compagnie": next_id,
                "Nom_Compagnie": nom,
                "Groupe": groupe,
                "Is_Marché": False
            }])], ignore_index=True)
            next_id += 1
    if "Marché" not in dim_compagnie["Nom_Compagnie"].values:
        dim_compagnie = pd.concat([dim_compagnie, pd.DataFrame([{
            "id_compagnie": 0, "Nom_Compagnie": "Marché", "Groupe": "Marché", "Is_Marché": True
        }])], ignore_index=True)

    colonnes_branche = [col for col in df_ca.columns if col not in [
        "Compagnie d'assurance", "TOTAL (AFF. DIRECTES)", "ACCEPTATIONS", "TOTAL (AFF. DIR & ACC)"
    ]]
    next_branche_id = 1 if dim_branche.empty else int(dim_branche["id_branche"].max()) + 1
    if 0 not in dim_branche["id_branche"].astype(int).values:
        dim_branche = pd.concat([dim_branche, pd.DataFrame([{
            "id_branche": 0, "Nom_Branche": "Toutes_Branches", "Type_Assurance": ""
        }])], ignore_index=True)
    for col in colonnes_branche:
        nom_branche = normaliser_nom_branche(col)
        if nom_branche not in dim_branche["Nom_Branche"].values:
            type_ass = "Vie" if "vie" in col.lower() else "Non-Vie"
            dim_branche = pd.concat([dim_branche, pd.DataFrame([{
                "id_branche": next_branche_id,
                "Nom_Branche": nom_branche,
                "Type_Assurance": type_ass
            }])], ignore_index=True)
            next_branche_id += 1

    if faits.empty:
        faits = pd.DataFrame(columns=[
            "id_fait", "id_temps", "id_compagnie", "id_branche", "primes_nettes",
            "resultat_technique", "sinistres_regles", "primes_cedees",
            "provisions_techniques", "resultat_net", "capitaux_propres",
            "primes_acquises", "charges_sinistres", "charges_acquisition_gestion_nettes"
        ])
    id_fait = 1 if faits.empty else int(faits["id_fait"].max()) + 1

    for _, ligne in df_ca.iterrows():
        nom_comp = ligne["Compagnie d'assurance"]
        nom_comp_ind = trouver_compagnie_equivalente(nom_comp, compagnies)
        if not nom_comp_ind:
            continue
        id_comp = get_id(dim_compagnie, "Nom_Compagnie", nom_comp_ind, "id_compagnie")

        for col in colonnes_branche:
            nom_branche = normaliser_nom_branche(col)
            id_branche = get_id(dim_branche, "Nom_Branche", nom_branche, "id_branche")

            primes = nettoyer_valeur(ligne[col])
            valeur_rt = 0
            valeur_sr = 0
            try:
                valeur_rt = nettoyer_valeur(df_res.loc[
                    df_res[df_res.columns[0]].apply(standardiser_nom) == standardiser_nom(nom_comp), col
                ].values[0])
            except:
                pass
            try:
                valeur_sr = nettoyer_valeur(df_sin.loc[
                    df_sin[df_sin.columns[0]].apply(standardiser_nom) == standardiser_nom(nom_comp), col
                ].values[0])
            except:
                pass

            primes_ced = 0
            provisions = 0
            res_net = 0
            cap_soc = 0
            if id_branche == 0:
                ligne_ind = df_ind[df_ind["COMPAGNIES"] == nom_comp_ind]
                if not ligne_ind.empty:
                    primes_ced = nettoyer_valeur(ligne_ind["PRIMES CEDEES"].values[0]) * 1_000_000
                    provisions = nettoyer_valeur(ligne_ind["PROVISIONS TECHNIQUES (1)"].values[0]) * 1_000_000
                    res_net = nettoyer_valeur(ligne_ind["RESULTATS COMPTABLES"].values[0]) * 1_000_000
                    cap_soc = nettoyer_valeur(ligne_ind["FONDS PROPRES (2)"].values[0]) * 1_000_000

            faits = pd.concat([faits, pd.DataFrame([{
                "id_fait": id_fait, "id_temps": id_temps,
                "id_compagnie": id_comp, "id_branche": id_branche,
                "primes_nettes": primes, "resultat_technique": valeur_rt,
                "sinistres_regles": valeur_sr, "primes_cedees": primes_ced,
                "provisions_techniques": provisions, "resultat_net": res_net,
                "capitaux_propres": cap_soc, "primes_acquises": 0,
                "charges_sinistres": 0, "charges_acquisition_gestion_nettes": 0
            }])], ignore_index=True)
            id_fait += 1
    
    # Ajout manuel des valeurs pour id_branche = 0 à partir du fichier indicateurs
    for nom_comp in df_ca["Compagnie d'assurance"].dropna().unique():
        nom_comp_ind = trouver_compagnie_equivalente(nom_comp, compagnies)
        if not nom_comp_ind:
            continue
        id_comp = get_id(dim_compagnie, "Nom_Compagnie", nom_comp_ind, "id_compagnie")
        id_branche = 0  # Toutes_Branches

        ligne_ind = df_ind[df_ind["COMPAGNIES"] == nom_comp_ind]
        if ligne_ind.empty:
            continue

        primes_ced = nettoyer_valeur(ligne_ind["PRIMES CEDEES"].values[0]) * 1_000_000
        provisions = nettoyer_valeur(ligne_ind["PROVISIONS TECHNIQUES (1)"].values[0]) * 1_000_000
        res_net = nettoyer_valeur(ligne_ind["RESULTATS COMPTABLES"].values[0]) * 1_000_000
        cap_soc = nettoyer_valeur(ligne_ind["FONDS PROPRES (2)"].values[0]) * 1_000_000

        faits = pd.concat([faits, pd.DataFrame([{
            "id_fait": id_fait, "id_temps": id_temps,
            "id_compagnie": id_comp, "id_branche": id_branche,
            "primes_nettes": 0, "resultat_technique": 0,
            "sinistres_regles": 0, "primes_cedees": primes_ced,
            "provisions_techniques": provisions, "resultat_net": res_net,
            "capitaux_propres": cap_soc, "primes_acquises": 0,
            "charges_sinistres": 0, "charges_acquisition_gestion_nettes": 0
        }])], ignore_index=True)
        id_fait += 1


    ligne_acquises = df_cpt[df_cpt[df_cpt.columns[0]].astype(str).str.strip().str.lower() == "1 primes acquises"]
    ligne_prestations = df_cpt[df_cpt[df_cpt.columns[0]].astype(str).str.strip().str.lower() == "4 charges de prestations"]
    ligne_gestion = df_cpt[df_cpt[df_cpt.columns[0]].astype(str).str.strip().str.lower() == "12 charges d'acquisition et de gestion nettes"]

    for col in colonnes_branche:
        nom_branche = normaliser_nom_branche(col)
        id_branche = get_id(dim_branche, "Nom_Branche", nom_branche, "id_branche")
        col_exists = col in df_cpt.columns

        primes = nettoyer_valeur(ligne_acquises[col].values[0]) if (not ligne_acquises.empty and col_exists) else 0
        prestations = nettoyer_valeur(ligne_prestations[col].values[0]) if (not ligne_prestations.empty and col_exists) else 0
        gestion = nettoyer_valeur(ligne_gestion[col].values[0]) if (not ligne_gestion.empty and col_exists) else 0

        faits = pd.concat([faits, pd.DataFrame([{
            "id_fait": id_fait, "id_temps": id_temps,
            "id_compagnie": 0, "id_branche": id_branche,
            "primes_nettes": 0, "resultat_technique": 0,
            "sinistres_regles": 0, "primes_cedees": 0,
            "provisions_techniques": 0, "resultat_net": 0,
            "capitaux_propres": 0, "primes_acquises": primes,
            "charges_sinistres": prestations,
            "charges_acquisition_gestion_nettes": gestion
        }])], ignore_index=True)
        id_fait += 1

    def enregistrer(df, nom):
        if nom not in wb.sheetnames:
            wb.create_sheet(nom)
        ws = wb[nom]
        ws.delete_rows(1, ws.max_row)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

    enregistrer(dim_compagnie, "dim_compagnie")
    enregistrer(dim_branche, "dim_branche")
    enregistrer(dim_temps, "dim_temps")
    enregistrer(faits, "Faits_KPI_Assurance")
    log_df = pd.concat([log_df, pd.DataFrame([{"Annee": int(annee), "NomFichier": fichier_IND}])], ignore_index=True)
    wb.save(FICHIER_TEMPLATE)
    log_df.to_excel(FICHIER_LOG, index=False)

# ====== AJOUT VALEURS POUR MARCHE GLOBAL (id_compagnie=0, id_branche=0) ======
ligne_ind_marche = df_ind[df_ind["COMPAGNIES"].astype(str).str.upper().str.contains("TOTAL")]
if not ligne_ind_marche.empty:
    # Get market totals from all files
    primes_cedees = nettoyer_valeur(ligne_ind_marche["PRIMES CEDEES"].values[0]) * 1_000_000
    provisions = nettoyer_valeur(ligne_ind_marche["PROVISIONS TECHNIQUES (1)"].values[0]) * 1_000_000
    resultat_net = nettoyer_valeur(ligne_ind_marche["RESULTATS COMPTABLES"].values[0]) * 1_000_000
    capitaux_propres = nettoyer_valeur(ligne_ind_marche["FONDS PROPRES (2)"].values[0]) * 1_000_000
    
    # Get resultat_technique from RESULTAT TECHNIQUE file
    rt_total = nettoyer_valeur(df_res[df_res[df_res.columns[0]].astype(str).str.upper().str.contains("TOTAL")][next((col for col in df_res.columns if "TOTAL" in col.upper()), None)].values[0])
    
    # Get sinistres_regles from SINISTRES REGLES file
    sr_total = nettoyer_valeur(df_sin[df_sin[df_sin.columns[0]].astype(str).str.upper().str.contains("TOTAL")][next((col for col in df_sin.columns if "TOTAL" in col.upper()), None)].values[0])
    
    # Get primes_nettes from CHIFFRES D'AFFAIRES file
    primes_nettes = nettoyer_valeur(df_ca[df_ca[df_ca.columns[0]].astype(str).str.upper().str.contains("TOTAL")]["TOTAL (AFF. DIR & ACC)"].values[0])
    
    # Get compte d'exploitation values
    primes_acquises = nettoyer_valeur(ligne_acquises["TOTAL (AFF. DIR+ACC)"].values[0]) if not ligne_acquises.empty else 0
    charges_sinistres = nettoyer_valeur(ligne_prestations["TOTAL (AFF. DIR+ACC)"].values[0]) if not ligne_prestations.empty else 0
    charges_gestion = nettoyer_valeur(ligne_gestion["TOTAL (AFF. DIR+ACC)"].values[0]) if not ligne_gestion.empty else 0

    faits = pd.concat([faits, pd.DataFrame([{
        "id_fait": id_fait, "id_temps": id_temps,
        "id_compagnie": 0, "id_branche": 0,
        "primes_nettes": primes_nettes,
        "resultat_technique": rt_total,
        "sinistres_regles": sr_total,
        "primes_cedees": primes_cedees,
        "provisions_techniques": provisions,
        "resultat_net": resultat_net,
        "capitaux_propres": capitaux_propres,
        "primes_acquises": primes_acquises,
        "charges_sinistres": charges_sinistres,
        "charges_acquisition_gestion_nettes": charges_gestion
    }])], ignore_index=True)
    id_fait += 1

# ====== AJOUT VALEURS id_branche=0 POUR CHAQUE COMPAGNIE ======
for _, ligne in df_ca.iterrows():
    nom_comp = ligne["Compagnie d'assurance"]
    if pd.isna(nom_comp) or "TOTAL" in str(nom_comp).upper():
        continue
        
    nom_comp_ind = trouver_compagnie_equivalente(nom_comp, compagnies)
    if not nom_comp_ind:
        continue
        
    id_comp = get_id(dim_compagnie, "Nom_Compagnie", nom_comp_ind, "id_compagnie")
    
    # Get total values from CHIFFRES D'AFFAIRES
    primes_nettes = nettoyer_valeur(ligne["TOTAL (AFF. DIR & ACC)"])
    
    # Get resultat_technique from RESULTAT TECHNIQUE file
    try:
        rt = nettoyer_valeur(df_res.loc[
            df_res[df_res.columns[0]].apply(standardiser_nom) == standardiser_nom(nom_comp), 
            "TOTAL (AFF. DIR & ACC)"
        ].values[0])
    except:
        rt = 0
    
    # Get sinistres_regles from SINISTRES REGLES file
    try:
        sr = nettoyer_valeur(df_sin.loc[
            df_sin[df_sin.columns[0]].apply(standardiser_nom) == standardiser_nom(nom_comp), 
            "TOTAL (AFF. DIR & ACC)"
        ].values[0])
    except:
        sr = 0
    
    # Get company indicators from PRINCIPAUX INDICATEURS
    ligne_ind = df_ind[df_ind["COMPAGNIES"] == nom_comp_ind]
    primes_cedees = nettoyer_valeur(ligne_ind["PRIMES CEDEES"].values[0]) * 1_000_000 if not ligne_ind.empty else 0
    provisions = nettoyer_valeur(ligne_ind["PROVISIONS TECHNIQUES (1)"].values[0]) * 1_000_000 if not ligne_ind.empty else 0
    resultat_net = nettoyer_valeur(ligne_ind["RESULTATS COMPTABLES"].values[0]) * 1_000_000 if not ligne_ind.empty else 0
    capitaux_propres = nettoyer_valeur(ligne_ind["FONDS PROPRES (2)"].values[0]) * 1_000_000 if not ligne_ind.empty else 0

    # Check if this company-total combination already exists
    existing = faits[(faits["id_compagnie"] == id_comp) & (faits["id_branche"] == 0)]
    if not existing.empty:
        # Update existing row
        idx = existing.index[0]
        faits.at[idx, "primes_nettes"] = primes_nettes
        faits.at[idx, "resultat_technique"] = rt
        faits.at[idx, "sinistres_regles"] = sr
    else:
        # Add new row
        faits = pd.concat([faits, pd.DataFrame([{
            "id_fait": id_fait, "id_temps": id_temps,
            "id_compagnie": id_comp, "id_branche": 0,
            "primes_nettes": primes_nettes,
            "resultat_technique": rt,
            "sinistres_regles": sr,
            "primes_cedees": primes_cedees,
            "provisions_techniques": provisions,
            "resultat_net": resultat_net,
            "capitaux_propres": capitaux_propres,
            "primes_acquises": 0,  # Only for market
            "charges_sinistres": 0,  # Only for market
            "charges_acquisition_gestion_nettes": 0  # Only for market
        }])], ignore_index=True)
        id_fait += 1